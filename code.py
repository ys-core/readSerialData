#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 1/8/2020 3:41 PM
# @Author  : yonnsongLee@163.com
# @Site    : 
# @File    : app.py
# @Software: PyCharm


import array
import sys
import time
import serial
import serial.tools.list_ports
import threading
import pyqtgraph as pg
from tkinter import *
from openpyxl import *

serialPortList = []
for item in serial.tools.list_ports.comports():
    serialPortList.append(str(item)[0:str(item).find("-")-1])

def controlPanel():
    global buttonList, btnStart, btnPause, btnStop, allSampleNumberEntry, rangeFromEntry, rangeToEntry, resultAverageCurrentEntry
    root = Tk()
    root.title(" CONTROL PANEL")
    # root.iconbitmap('m.ico')

    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws / 2) - (800 / 2)
    y = (hs / 2) - (500 / 2)
    root.geometry('%dx%d+%d+%d' % (800, 500, x, y))
    labelTop = Label(root,text="1. Select the serial port",font=("Calibri",13))
    labelTop.place(x = 20,y = 10,anchor = NW)
    buttonList = StringVar()
    for index,port in enumerate(serialPortList):
        r = Radiobutton(root, variable=buttonList, value=port, text=port, bd=10, font=("Calibri",10))
        r.place(x=100*index+50, y=50)
    buttonList.set(serialPortList[0])
    if(len(serialPortList) > 0):
        labelButton = Label(root, text="2. Start acquisition",font=("Calibri",13))
        labelButton.place(x = 20,y = 100,anchor = NW)
        btnStart = Button(root, text=" START", bg="SeaGreen", fg="white",width=10, command=startAcquisition)
        btnPause = Button(root, text=" PAUSE", bg="Olive", fg="white", width=10, command=pauseAcquisition)
        btnStop = Button(root, text=" STOP ", bg="red", fg="white", width=10, command=stopAcquisition)
        btnStart.place(x=60, y=150)
        labelSampleNumberTitle = Label(root, text="3. Get the number of samples",font=("Calibri",13))
        labelSampleNumberTitle.place(x = 20, y = 200, anchor = NW)
        allSampleNumberLabel = Label(root,text="Sample Num:",font=("Calibri",10))
        allSampleNumberEntry = Entry(root,width=10,justify=CENTER)
        btnGetSampleNumber = Button(root, text="Get", bg="OrangeRed", fg="white", width=6,command=getSampleNumber)
        allSampleNumberLabel.place(x = 60, y = 250)
        allSampleNumberEntry.place(x = 175,y = 250)
        btnGetSampleNumber.place(x = 290, y = 245)
        labelAverageCurrentTitle = Label(root, text="4. Calculate the average current, view waveform and input the range to calculate",font=("Calibri",13))
        labelAverageCurrentTitle.place(x = 20, y = 300, anchor = NW)
        rangeFromLabel = Label(root,text="From",font=("Calibri",10))
        rangeFromEntry = Entry(root,width=10,justify=CENTER)
        rangeToLabel = Label(root,text="To",font=("Calibri",10))
        rangeToEntry = Entry(root,width=10,justify=CENTER)
        resultAverageCurrentLabel = Label(root,text="Average Value is :",font=("Calibri",10))
        resultAverageCurrentEntry = Entry(root,width=19,bd=3,fg="MediumBlue",justify=CENTER)
        btnAverageCurrent = Button(root, text="Average Value", bg="orange", fg="white",width=15, command=calculateAverageCurrent)
        rangeFromLabel.place(x = 60,y = 350)
        rangeFromEntry.place(x = 110,y = 350)
        rangeToLabel.place(x = 230,y = 350)
        rangeToEntry.place(x = 280,y = 350)
        # resultAverageCurrentLabel.place(x = 400, y = 350)
        resultAverageCurrentEntry.place(x = 570, y = 350)
        btnAverageCurrent.place(x = 400, y = 346)
        labelAverageCurrentTitle = Label(root,text="5. Save data to Excel file",font=("Calibri", 13))
        labelAverageCurrentTitle.place(x=20, y=410, anchor=NW)
        btnIntoExcel = Button(root, text="To Excel", bg="Tomato", fg="white",width=15, command=saveDataIntoExcel)
        btnIntoExcel.place(x = 300, y = 410)
    else:
        labelWarningNoSerialPortConnection = Label(root, text=" No serial port available, please connect.. ",fg="red",font=("Calibri", 20))
        labelWarningNoSerialPortConnection.place(x=100, y=300)
    root.mainloop()
def startAcquisition():
    global buttonList, mSerial, btnStart, btnStop,root
    portx = buttonList.get()
    mSerial = serial.Serial(portx, baudrate=115200, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE,stopbits=serial.STOPBITS_ONE, timeout=None)
    if (mSerial.isOpen()):
        mSerial.flushInput()
        btnStart.place_forget()
        btnStop.place(x=300, y=150)
        th1 = threading.Thread(target=Serial)  # watch and read the serialPort data
        th1.setDaemon(True)
        th1.start()
    else:
        print("open failed")
        mSerial.close()  # close serial port
def pauseAcquisition():
    global buttonList, mSerial, btnStart, btnStop, btnPause, pausing
    if (mSerial.isOpen):
        mSerial.close()
        btnPause.place_forget()
        btnStop.place_forget()
        btnStart.place(x=60, y=150)
    pausing = False
def stopAcquisition():
    global buttonList, mSerial, btnStart, btnPause, btnStop
    if(mSerial.isOpen):
        mSerial.close()
        btnPause.place_forget()
        btnStop.place_forget()
        btnStart.place(x=60, y=150)
        # allData.clear()
def setRangeFrom():
    print("from...")
def setRangeTo():
    print("to..")
def getSampleNumber():
    global allSampleNumberEntry
    allSampleNumberEntry.delete(0,'end')
    allSampleNumberEntry.insert(0,len(allData))
def calculateAverageCurrent():
    global rangeFromEntry, rangeToEntry, resultAverageCurrentEntry
    resultAverageCurrentEntry.delete(0,'end')
    try:
        From = int(rangeFromEntry.get())
        To = int(rangeToEntry.get())
    except Exception:
        resultAverageCurrentEntry.insert(0, 'input should be integer')
    else:
        From = From -1
        To = To - 1
        count = 0
        if(len(allData) <= 0):
            resultAverageCurrentEntry.insert(0, "No sample data")
        else:
            if (From < 0 or To < 0 or To < From or To>=len(allData)):
                resultAverageCurrentEntry.insert(0, "input valid")
            else:
                sampleNum = To - From + 1
                while(From <= To):
                    count += allData[From]
                    From += 1
                average_val = count / sampleNum
                resultAverageCurrentEntry.insert(0,average_val)
def saveDataIntoExcel():
    wb_name = str(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())).replace(':','_') + '.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'current_data'
    ws.cell(1,1,'Current:A')
    for i in range(len(allData)):
        ws.cell(i+2,1,allData[i])
    wb.save(wb_name)
def Serial():
    global mSerial
    while(True):
        _data = str(mSerial.readline())
        # _data = str(_data)[2:-5]
        # _data1 = str(_data).replace('\\n','')
        _data = _data.lstrip("b").strip("'").rstrip("\\r\\n")
        # print(_data)
        if(_data == ""):
            continue
        else:
            global i,startTime
            try:
                _data = float(_data)
            except Exception:
                print("Not float data")
            else:
                now = pg.ptime.time()
                i = i + 1
                if(i> 50000 or pausing):
                    break
                _data = float(_data)
                # print(i,_data)
                if i==1:
                    startTime = float(pg.ptime.time())
                    timeList.append(0)
                    allData.append(_data)
                else:
                    timeList.append(float(now)-startTime)
                    allData.append(_data)
def plotData():
    curve.setData(x=timeList,y=allData,pen="g")
    # curve.setData(allData)


if __name__ == "__main__":
    app = pg.mkQApp()
    win = pg.GraphicsWindow()
    win.setWindowTitle(u' real-time current curve..')
    win.resize(1200, 550)
    # p1 = win.addPlot(row=0, col=0)
    # p1.showGrid(x=True, y=True, alpha=0.1)
    # p.setRange(xRange=[0, 20000], padding=0)
    # p1.setLabel(axis='left', text='Current/A')
    # p1.setLabel(axis='bottom', text=' Number')
    # p1.setTitle('real-time current')
    # curve1 = p1.plot()
    p = win.addPlot(row=0,col=0)
    p.showGrid(x=True, y=True, alpha=0.1)
    p.setLabel(axis='left', text='Current/A')
    p.setLabel(axis='bottom', text=' Time(s)')
    p.setRange(xRange=[0, 5], yRange=[-10,9000], padding=0)
    p.setTitle('real-time current')
    curve = p.plot()


    allData = []
    timeList = []
    i = 0
    pausing = False
    th2 = threading.Thread(target = controlPanel)  # loading the control panel board
    th2.setDaemon(True)
    th2.start()
    # th1 = threading.Thread(target = Serial)  # watch and read the serialPort data
    # th1.setDaemon(True)
    # th1.start()
    timer = pg.QtCore.QTimer()
    timer.timeout.connect(plotData)
    timer.start(20)

    app.exec_()

    # th1.join()















